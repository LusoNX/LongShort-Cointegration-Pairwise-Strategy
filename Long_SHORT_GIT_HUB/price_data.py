from scipy.optimize import minimize
import numpy as np
import pandas as pd
import datetime as dt
import copy
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
import sklearn
import time
import math
import plotly.express as px
import plotly.graph_objects as go
from cvxpy import *
from scipy.optimize import nnls
#import scipy.optimize
from datetime import datetime
from scipy.optimize import nnls
import seaborn as sns
import urllib
from sqlalchemy import create_engine
import pyodbc
import yfinance as yf
import investpy
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, FORMULAE
from datetime import timedelta
import database_directory
conn_str = database_directory.main()
cnn_url = f"access+pyodbc:///?odbc_connect={urllib.parse.quote_plus(conn_str)}"
acc_engine = create_engine(cnn_url)



def download_bloomberg_tickers(hist_data,update_days_time):
    df_fund_index = pd.read_sql("SELECT * FROM AssetIndex ",acc_engine).sort_values("ID_INSTRUMENT")
    df_fund_index = df_fund_index[df_fund_index["Instrument_Type"] == "Index"]
    bb_ids = list(df_fund_index[df_fund_index["data_source"] == "bloomberg"]["ID_INSTRUMENT"].values)
    non_bb_ids =  list(df_fund_index[df_fund_index["data_source"] != "bloomberg"]["ID_INSTRUMENT"].values)
    old_bb_ids = list(pd.read_sql("SELECT ID_INSTRUMENT FROM PriceIndex",acc_engine).sort_values("ID_INSTRUMENT")["ID_INSTRUMENT"].unique())
    old_bb_ids = list(set(old_bb_ids).symmetric_difference(set(non_bb_ids)))
    new_bb_ids = list(set(old_bb_ids).symmetric_difference(set(bb_ids)))

    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = "ISIN"
    ws1["A2"] = "BB_Ticker"

    if hist_data == True:
        start_date = "01/01/2000"
        df_fund_index = df_fund_index[np.isin(df_fund_index, new_bb_ids).any(axis=1)]
        days_ret = 5900
    else:
        start_date = datetime.today() - timedelta(days=update_days_time) # Last 8 days
        start_date = start_date.strftime("%d/%m/%Y")
        df_fund_index = df_fund_index[np.isin(df_fund_index, old_bb_ids).any(axis=1)]
        days_ret = update_days_time

    for i in range(0,len(df_fund_index["ID_INSTRUMENT"])):
        column_number = i+2
        letter_columns = get_column_letter(column_number)
        row_value_1 =letter_columns +str(1)
        row_value_2 = letter_columns + str(2)
        row_value_3 = letter_columns +str(3)
        id_instrument = df_fund_index["ID_INSTRUMENT"].iloc[i]
        df_price_exists = pd.read_sql("SELECT * FROM PriceIndex WHERE ID_INSTRUMENT ={}".format(id_instrument),acc_engine)
        ws1[row_value_1] = df_fund_index["code"].iloc[i]
        ws1[row_value_2] =df_fund_index["code"].iloc[i]
        if row_value_3 == "B3":
            ws1["A3"] = 'replace_ti=BDH({};"PX_LAST";"{}";"";"Dir=V";"CDR=5D";"CshAdjNormal=Y";"CshAdjAbnormal=Y";"CapChg=Y";"Days=A";"Dts=S";"FX=USD";"cols=1;rows={}")'.format(row_value_2,start_date,days_ret)
        else:
            ws1[row_value_3] = 'replace_ti=BDH({};"PX_LAST";"{}";"";"Dir=V";"CDR=5D";"CshAdjNormal=Y";"CshAdjAbnormal=Y";"CapChg=Y";"Days=A";"Dts=H";"FX=USD";"cols=1;rows={}")'.format(row_value_2,start_date,days_ret)

    if hist_data == True:
        wb.save(filename =r"P:\Public\LuanF\L_S_Strat\BB_import\bb_import_new_data.xlsx")
    else:
        wb.save(filename =r"P:\Public\LuanF\L_S_Strat\BB_import\bb_import_old_data.xlsx")

    pass


def yahoo_finance_data(id_instrument):
    asset_index = pd.read_sql("SELECT * FROM AssetIndex WHERE ID_INSTRUMENT ={}".format(id_instrument),acc_engine)
    id_instrument = asset_index["ID_INSTRUMENT"].values[0]
    ticker = asset_index["code"].values[0]
    df_price_exists = pd.read_sql("SELECT * FROM PriceIndex WHERE ID_INSTRUMENT = {}".format(id_instrument),acc_engine).sort_values("Data")

    if df_price_exists.empty:
        current_date = datetime.now().strftime("%Y-%m-%d")
        df_price = yf.download(ticker,'2008-1-1', current_date)

        df_price = df_price.astype(float)
        df_price["ID_INSTRUMENT"] = id_instrument
        df_price.reset_index(inplace = True)
        df_price.rename(columns = {"Date":"Data"},inplace = True)
        df_price.rename(columns = {"Adj Close":"Adj_Close"},inplace = True)
        df_price.set_index("ID_INSTRUMENT",inplace = True)
        df_price.to_sql("PriceIndex",acc_engine,if_exists = "append")
        print("New price Data for ID: {} ".format(id_instrument))
    else:
        last_price_date = str((df_price_exists.iloc[-1]["Data"]).strftime("%Y-%m-%d"))
        current_date = str(datetime.now().strftime("%Y-%m-%d"))
        
        if current_date == last_price_date:
            print("Data is already up to date")
            pass
        else:

            df_price = yf.download(ticker,last_price_date, current_date)
            df_price = df_price.iloc[1::]
            df_price = df_price.astype(float)
            df_price["ID_INSTRUMENT"] = id_instrument
            df_price.reset_index(inplace = True)
            df_price.rename(columns = {"Date":"Data"},inplace = True)
            df_price.rename(columns = {"Adj Close":"Adj_Close"},inplace = True)
            df_price.set_index("ID_INSTRUMENT",inplace = True)
            df_price.to_sql("PriceIndex",acc_engine,if_exists = "append")
            print("Data for ID: {}  was updated".format(id_instrument))




def get_data():
    unique_ids = pd.read_sql("SELECT ID_INSTRUMENT,data_source FROM AssetIndex",acc_engine)
    unique_ids = unique_ids[unique_ids["data_source"] != "bloomberg"]
    #unique_ids = unique_ids[unique_ids["ID_INSTRUMENT"]==10]
    for i,x in zip(unique_ids["ID_INSTRUMENT"],unique_ids["data_source"]):
        yahoo_finance_data(i)
        


def main():
    get_data()


main()